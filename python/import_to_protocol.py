#!/usr/bin/env python3
"""
import_to_protocol.py — Bulk-import a data file into a CDD Vault protocol via
the Slurps API (the programmatic equivalent of the web UI's "Import Data").

HARD RULE: this script NEVER reads or prints the contents of the data file.
The file is streamed to CDD as a binary handle (multipart upload); the only
thing that crosses to *this* process's stdout is metadata — slurp id, state,
protocol/readout ids and names, status codes, and error/warning counts.

CDD Slurps flow (async):
  1. POST /vaults/{v}/slurps    multipart: a `file` part + a `json` part with
                                {project, mapping_template, runs, autoreject}.
                                Response: {"id": <slurp_id>, ...}.
  2. GET  /vaults/{v}/slurps/{slurp_id}   poll until `state` is terminal
                                ("committed" = success; "rejected"/"invalid"/
                                "canceled" = failure).

Mapping (file columns -> protocol readouts) is resolved one of two ways:
  * --mapping-template NAME   reference a template created once in the web UI
                              (the confirmed, reliable path).
  * --mapping-json PATH       a JSON file whose contents are merged verbatim
                              into the slurp payload (inline mappings). The
                              inline-mapping schema is NOT publicly documented,
                              so this is a pass-through — you own the JSON.

Use --list-protocols / --describe-protocol to discover protocol + readout ids
(metadata only, writes nothing) before building a mapping.

Per-protocol column->readout mappings live in config/config.yaml. With
--protocol set, the upload header is validated against that config (column
names only, never data values) — as a standalone --check-mapping pass or as an
automatic pre-flight that aborts a real import before any column mismatch
reaches CDD.

CLI:
  # discover
  python import_to_protocol.py --vault 7108 --list-protocols
  python import_to_protocol.py --vault 7108 --describe-protocol 12345

  # offline mapping check (no token/network)
  python import_to_protocol.py --file data.csv --protocol 12345 --check-mapping

  # import (dry-run first — validates + prints payload, posts nothing)
  python import_to_protocol.py --vault 7108 --file data.csv \
      --project "MyProject" --mapping-template "MDR1 upload" --dry-run
  # ...then drop --dry-run to actually submit + poll
"""

import argparse
import csv
import json
import sys
import time
from pathlib import Path

# Reuse the auth + base-url primitives from the exporter (same vault, same token).
from get_library import API_BASE, load_token, make_session

# Slurp lifecycle states. Anything not in STARTED is terminal.
TERMINAL_STATES = {"committed", "rejected", "invalid", "canceled"}
SUCCESS_STATE = "committed"

# Content types CDD accepts for a slurp, keyed by file extension.
MIME_BY_EXT = {
    ".csv": "text/csv",
    ".sdf": "chemical/x-mdl-sdfile",
    ".zip": "application/zip",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}


# ---------- discovery (metadata only) ----------

def _objects(data):
    """CDD list endpoints return {'objects': [...]} or a bare list."""
    if isinstance(data, dict) and "objects" in data:
        return data["objects"]
    return data if isinstance(data, list) else []


def list_protocols(session, vault):
    """Print id + name + readout count for every protocol in the vault."""
    r = session.get(f"{API_BASE}/vaults/{vault}/protocols", timeout=60)
    if r.status_code != 200:
        sys.exit(f"ERR list-protocols status={r.status_code} body_len={len(r.content)}")
    protos = _objects(r.json())
    print(f"=== protocols in vault {vault}: {len(protos)} ===")
    for p in protos:
        rds = p.get("readout_definitions") or []
        print(f"  id={p.get('id')}  name={p.get('name')!r}  readouts={len(rds)}")


def describe_protocol(session, vault, pid):
    """Print the readout definitions (id, name, data type) for one protocol.

    These ids/names are what a mapping template (or inline mapping) targets.
    Falls back to dumping top-level keys if the readout shape differs from the
    expected `readout_definitions` list.
    """
    r = session.get(f"{API_BASE}/vaults/{vault}/protocols/{pid}", timeout=60)
    if r.status_code != 200:
        sys.exit(f"ERR describe-protocol status={r.status_code} body_len={len(r.content)}")
    p = r.json()
    print(f"=== protocol id={p.get('id')} name={p.get('name')!r} ===")
    rds = p.get("readout_definitions")
    if not isinstance(rds, list):
        print(f"WARN: no 'readout_definitions' list; top-level keys={sorted(p)}")
        return
    print(f"readout_definitions={len(rds)}:")
    for rd in rds:
        dt = rd.get("data_type")
        dt = dt.get("name") if isinstance(dt, dict) else dt
        print(f"  id={rd.get('id')}  name={rd.get('name')!r}  type={dt}")


# ---------- config + mapping validation ----------

def load_protocol_mapping(config_path, pid):
    """Read config/config.yaml and return the block for protocol `pid`.

    Returns a dict with 'name', 'identifiers', 'readouts' (column -> readout id).
    Exits if the file or the protocol block is missing.
    """
    import yaml  # lazy — only needed when a config-driven mode is used

    p = Path(config_path)
    if not p.exists():
        sys.exit(f"ERR: config not found: {p}")
    cfg = yaml.safe_load(p.read_text()) or {}
    protos = cfg.get("protocols") or {}
    block = protos.get(pid) or protos.get(str(pid))
    if block is None:
        sys.exit(f"ERR: protocol {pid} not in {p}; "
                 f"known={sorted(protos)}")
    return block


def read_header(file_path):
    """Return the upload file's column names ONLY (schema, never data rows).

    CSV: first row. XLSX: first row of the first sheet. The BOM some exporters
    prepend to the first header cell is stripped.
    """
    p = Path(file_path)
    ext = p.suffix.lower()
    if ext in (".csv", ".txt"):
        with open(p, newline="", encoding="utf-8-sig") as fh:
            header = next(csv.reader(fh), [])
    elif ext == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(p, read_only=True)
        ws = wb.active
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        wb.close()
    else:
        sys.exit(f"ERR: --check-mapping supports .csv/.xlsx, not {ext!r}")
    return [str(h).lstrip("﻿").strip() for h in header if h not in (None, "")]


def _norm(s):
    """Collapse whitespace runs — incl. the CR/LF in multi-line CRO headers like
    'CLint (raw)\\r\\r\\n(µL/min/mg)' — so config keys can be clean single-line."""
    return " ".join(str(s).split()) if s is not None else ""


def validate_header(header, block):
    """Compare upload columns against a protocol's config block.

    Matching is whitespace-insensitive (see _norm) so multi-line headers match
    clean single-line config keys. Returns original column names for display.

    Returns (mapped, ignored, unmapped_cols, missing_readouts):
      - mapped: file columns that match a readout or an identifier
      - ignored: file columns in the config's `ignore` list (present on purpose,
        not imported — e.g. a redundant 'Molecule Name' next to the batch id)
      - unmapped_cols: file columns with no home in the config (likely typos/strays)
      - missing_readouts: config readouts not present in the file
    """
    readouts = block.get("readouts") or {}
    idents = set((block.get("identifiers") or {}).values())
    ignore = set(block.get("ignore") or [])
    importable_n = {_norm(k) for k in readouts} | {_norm(v) for v in idents}
    ignore_n = {_norm(x) for x in ignore}
    cols = list(dict.fromkeys(header))  # de-dup, preserve order (original names)
    col_norms = {_norm(c) for c in cols}
    mapped = [c for c in cols if _norm(c) in importable_n]
    ignored = [c for c in cols if _norm(c) in ignore_n and _norm(c) not in importable_n]
    unmapped = [c for c in cols if _norm(c) not in importable_n and _norm(c) not in ignore_n]
    missing = [r for r in readouts if _norm(r) not in col_norms]
    return mapped, ignored, unmapped, missing


def count_rows_missing_identifiers(file_path, block):
    """Count data rows whose identifier columns are ALL blank (CDD rejects these).

    Reads identifier-column VALUES only to count blanks — values are never
    printed. A row with at least one identifier filled is fine (CDD links the
    molecule from the batch id). Returns (total_rows, missing_rows); (None, None)
    if the format isn't row-readable or no identifier column is present.
    """
    idents = list((block.get("identifiers") or {}).values())
    p = Path(file_path)
    ext = p.suffix.lower()
    if not idents:
        return (None, None)
    if ext in (".csv", ".txt"):
        rows = list(csv.reader(open(p, newline="", encoding="utf-8-sig")))
    elif ext == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(p, read_only=True)
        rows = [[c.value for c in r] for r in wb.active.iter_rows()]
        wb.close()
    else:
        return (None, None)
    if not rows:
        return (0, 0)
    hdr = [str(h).lstrip("﻿").strip() if h is not None else "" for h in rows[0]]
    idx = [hdr.index(c) for c in idents if c in hdr]
    if not idx:
        return (None, None)
    blank = lambda v: v is None or str(v).strip() == ""
    total = missing = 0
    for r in rows[1:]:
        if all(blank(c) for c in r):  # skip wholly-empty lines
            continue
        total += 1
        if all(i >= len(r) or blank(r[i]) for i in idx):
            missing += 1
    return (total, missing)


def write_cleaned_csv(src_path, block, dst_path):
    """Copy a CSV to dst, dropping rows whose identifier columns are all blank.

    Removes the junk rows the prep tool stamps with Run Lab/Provider Name (and
    any wholly-empty lines). Values are copied verbatim, never printed. Returns
    (kept, dropped).
    """
    idents = list((block.get("identifiers") or {}).values())
    rows = list(csv.reader(open(src_path, newline="", encoding="utf-8-sig")))
    if not rows:
        return (0, 0)
    hdr = [str(h).lstrip("﻿").strip() if h is not None else "" for h in rows[0]]
    idx = [hdr.index(c) for c in idents if c in hdr]
    blank = lambda v: v is None or str(v).strip() == ""
    kept = [rows[0]]
    dropped = 0
    for r in rows[1:]:
        empty_line = all(blank(c) for c in r)
        no_id = bool(idx) and all(i >= len(r) or blank(r[i]) for i in idx)
        if empty_line or no_id:
            dropped += 1
        else:
            kept.append(r)
    with open(dst_path, "w", newline="", encoding="utf-8") as fh:
        csv.writer(fh).writerows(kept)
    return (len(kept) - 1, dropped)


def check_mapping(file_path, config_path, pid):
    """Offline pre-flight: header lines up with config + no empty-identifier rows.

    Header check reads column names only; the row check reads identifier-column
    values to COUNT blanks (never prints them). No network."""
    block = load_protocol_mapping(config_path, pid)
    header = read_header(file_path)
    mapped, ignored, unmapped, missing = validate_header(header, block)
    total_rows, missing_id_rows = count_rows_missing_identifiers(file_path, block)
    print(f"=== mapping check: protocol={pid} name={block.get('name')!r} ===")
    print(f"file={Path(file_path).name}  file_columns={len(header)}  "
          f"config_readouts={len(block.get('readouts') or {})}"
          + (f"  data_rows={total_rows}" if total_rows is not None else ""))
    print(f"mapped={len(mapped)}")
    if ignored:
        print(f"ignored ({len(ignored)}) — present but intentionally not imported:")
        for c in ignored:
            print(f"  - {c!r}")
    if unmapped:
        print(f"UNMAPPED file columns ({len(unmapped)}) — no readout/identifier "
              f"in config, will not import:")
        for c in unmapped:
            print(f"  - {c!r}")
    if missing:
        print(f"MISSING readouts ({len(missing)}) — in config but not in file "
              f"(left blank on import):")
        for r in missing:
            print(f"  - {r!r}")
    if missing_id_rows:
        print(f"EMPTY-IDENTIFIER rows ({missing_id_rows}/{total_rows}) — no "
              f"Molecule Name / Batch ID; CDD REJECTS these. Remove them from "
              f"the file before importing.")
    ok = not unmapped and not missing_id_rows
    problems = []
    if unmapped:
        problems.append("UNMAPPED columns")
    if missing_id_rows:
        problems.append("EMPTY-IDENTIFIER rows")
    print(f"\nmapping_ok={ok}"
          + ("" if ok else f"  (resolve {' + '.join(problems)} first)"))
    return ok


# ---------- payload ----------

def build_payload(project, mapping_template=None, mapping_json=None,
                  autoreject=True, runs=None):
    """Assemble the slurp `json` part. mapping_json (a dict) is merged verbatim."""
    payload = {"project": project, "autoreject": autoreject}
    if mapping_template:
        payload["mapping_template"] = mapping_template
    if runs:
        payload["runs"] = runs
    if mapping_json:
        # Inline-mapping schema is user-supplied and unverified — pass through.
        payload.update(mapping_json)
    return payload


def _build_runs(args):
    """Collect the optional run metadata into CDD's `runs` dict, if any given."""
    run = {
        k: v for k, v in (
            ("run_date", args.run_date),
            ("place", args.run_place),
            ("person", args.run_person),
            ("conditions", args.run_conditions),
        ) if v is not None
    }
    return run or None


# ---------- submit + poll ----------

def submit_slurp(session, vault, file_path, payload):
    """POST the file + payload, return the new slurp id. File is streamed, not read."""
    p = Path(file_path)
    if not p.exists():
        sys.exit(f"ERR: file not found: {p}")
    mime = MIME_BY_EXT.get(p.suffix.lower(), "application/octet-stream")
    print(f"submitting file={p.name} bytes={p.stat().st_size} mime={mime}")
    print(f"payload_keys={sorted(payload)} project={payload.get('project')!r} "
          f"mapping_template={payload.get('mapping_template')!r} "
          f"autoreject={payload.get('autoreject')}")
    with open(p, "rb") as fh:
        r = session.post(
            f"{API_BASE}/vaults/{vault}/slurps",
            files={"file": (p.name, fh, mime)},
            data={"json": json.dumps(payload)},
            timeout=300,
        )
    if r.status_code not in (200, 201):
        # Auth / endpoint errors carry an API message, not row data — safe to show.
        # 400/422 (validation) may echo column values, so stay metadata-only there.
        detail = f" body={r.text[:300]!r}" if r.status_code in (401, 403, 404) else ""
        sys.exit(f"ERR submit status={r.status_code} "
                 f"body_len={len(r.content)}{detail}")
    obj = r.json()
    sid = obj.get("id")
    if sid is None:
        sys.exit(f"ERR submit: response had no 'id'; keys={sorted(obj)}")
    print(f"slurp_id={sid}")
    return sid


def poll_slurp(session, vault, slurp_id, interval=5.0, timeout=600.0):
    """Poll the slurp until a terminal state or timeout. Return the final object."""
    url = f"{API_BASE}/vaults/{vault}/slurps/{slurp_id}"
    waited = 0.0
    while True:
        r = session.get(url, timeout=60)
        obj = r.json() if r.status_code == 200 else {}
        state = obj.get("state", f"http_{r.status_code}")
        print(f"  slurp={slurp_id} state={state} waited={waited:.0f}s")
        if state in TERMINAL_STATES:
            return obj
        if waited >= timeout:
            print(f"WARN: poll timeout after {waited:.0f}s; last state={state}")
            return obj
        time.sleep(interval)
        waited += interval


def _shape(obj, depth=0):
    """Describe a JSON value's STRUCTURE without leaking string values.

    Scalars: numbers/bools shown as-is, strings as <str:len>. Dicts: keys with
    nested shapes. Lists: length + the shape of the first element. Lets us see
    where error detail lives in a slurp object without printing row data.
    """
    pad = "  " * depth
    if isinstance(obj, dict):
        lines = []
        for k, v in obj.items():
            head = _shape(v, depth + 1)
            lines.append(f"{pad}  {k}: {head}")
        return "{\n" + "\n".join(lines) + f"\n{pad}}}"
    if isinstance(obj, list):
        inner = _shape(obj[0], depth + 1) if obj else "empty"
        return f"[{len(obj)}] of {inner}"
    if isinstance(obj, str):
        return f"<str:{len(obj)}>"
    return repr(obj)  # numbers, bools, None — safe


def report_outcome(obj):
    """Summarise a finished slurp. Counts only — error bodies may echo row data.

    Field names confirmed from a real committed slurp (CDD vault 7108):
    records_committed / records_processed / total_records / import_errors /
    import_warnings.
    """
    state = obj.get("state")
    ok = state == SUCCESS_STATE
    print(f"\nfinal_state={state}  success={ok}")
    for key in ("records_committed", "records_processed", "total_records"):
        if key in obj:
            print(f"  {key}={obj[key]}")
    for key in ("import_errors", "import_warnings"):
        v = obj.get(key)
        if isinstance(v, list):
            if v:
                print(f"  {key}={len(v)}")
        elif v not in (None, "", 0):
            print(f"  {key}={v}")
    if not ok:
        print("NOTE: inspect the slurp in the CDD web UI for per-row reasons, "
              "or run `--show-slurp <id>` to see its structure (keys/counts, no "
              "data values).")
    return ok


# ---------- main ----------

def main():
    ap = argparse.ArgumentParser(
        description="Bulk-import a data file into a CDD Vault protocol via the "
                    "Slurps API (file is streamed to CDD; stdout is metadata only).",
    )
    ap.add_argument("--vault", type=int, default=7108)
    ap.add_argument("--token-file", default="~/.cdd_token")

    # discovery modes
    ap.add_argument("--list-protocols", action="store_true",
                    help="List protocols (id, name, readout count), then exit.")
    ap.add_argument("--describe-protocol", type=int, metavar="PID",
                    help="Print a protocol's readout definitions, then exit.")
    ap.add_argument("--show-slurp", type=int, metavar="SLURP_ID",
                    help="Fetch a slurp and print its structure (keys/counts/"
                         "shape, no data values), then exit. For diagnosing a "
                         "rejected import.")

    # config-driven mapping
    ap.add_argument("--config", default="config/config.yaml",
                    help="YAML of per-protocol mappings (default config/config.yaml).")
    ap.add_argument("--protocol", type=int, metavar="PID",
                    help="Protocol id; used to look up its mapping block in --config.")
    ap.add_argument("--check-mapping", action="store_true",
                    help="Offline: validate the --file header against the "
                         "--protocol mapping in --config, then exit (no network).")

    # import args
    ap.add_argument("--file", help="Data file to import (.csv/.sdf/.zip/.xlsx).")
    ap.add_argument("--project", help="CDD project name (required for import).")
    ap.add_argument("--mapping-template",
                    help="Name of an existing mapping template (web UI).")
    ap.add_argument("--mapping-json",
                    help="Path to a JSON file merged verbatim into the payload "
                         "(inline mappings — schema is your responsibility).")
    ap.add_argument("--no-autoreject", action="store_true",
                    help="Allow commit despite suspicious events/errors "
                         "(autoreject defaults ON — safer).")
    ap.add_argument("--drop-empty-rows", action="store_true",
                    help="Drop rows whose identifier columns are all blank (junk "
                         "rows the prep tool stamps with Run Lab/Provider Name) "
                         "before upload. Writes a cleaned temp copy; original "
                         "untouched. Needs --protocol; CSV only.")
    # optional run metadata
    ap.add_argument("--run-date")
    ap.add_argument("--run-place")
    ap.add_argument("--run-person")
    ap.add_argument("--run-conditions")
    # polling
    ap.add_argument("--poll-interval", type=float, default=5.0)
    ap.add_argument("--poll-timeout", type=float, default=600.0)
    ap.add_argument("--dry-run", action="store_true",
                    help="Validate inputs and print the payload; post nothing.")
    args = ap.parse_args()

    # Offline mode — no token / network needed.
    if args.check_mapping:
        if not args.file or args.protocol is None:
            sys.exit("ERR: --check-mapping needs --file and --protocol")
        ok = check_mapping(args.file, args.config, args.protocol)
        sys.exit(0 if ok else 1)

    token = load_token(args.token_file)
    session = make_session(token)

    if args.list_protocols:
        list_protocols(session, args.vault)
        return
    if args.describe_protocol is not None:
        describe_protocol(session, args.vault, args.describe_protocol)
        return
    if args.show_slurp is not None:
        r = session.get(f"{API_BASE}/vaults/{args.vault}/slurps/{args.show_slurp}",
                        timeout=60)
        if r.status_code != 200:
            sys.exit(f"ERR show-slurp status={r.status_code}")
        print(f"=== slurp {args.show_slurp} structure (no data values) ===")
        print(_shape(r.json()))
        return

    # When a protocol is named, fall back to its config block for project /
    # mapping_template if not passed on the CLI (CLI always wins).
    project = args.project
    mapping_template = args.mapping_template
    if args.protocol is not None:
        block = load_protocol_mapping(args.config, args.protocol)
        project = project or block.get("project")
        mapping_template = mapping_template or (block.get("mapping_template") or None)

    if not args.file or not project:
        sys.exit("ERR: import needs --file and a project "
                 "(--project, or `project:` in the config block)")
    if not mapping_template and not args.mapping_json:
        sys.exit("ERR: supply --mapping-template NAME, set `mapping_template:` "
                 "in the config block, or use --mapping-json PATH")

    # Optionally strip junk rows into a cleaned temp copy; original untouched.
    work_file = args.file
    if args.drop_empty_rows:
        if args.protocol is None:
            sys.exit("ERR: --drop-empty-rows needs --protocol (to know identifiers)")
        if Path(args.file).suffix.lower() not in (".csv", ".txt"):
            sys.exit("ERR: --drop-empty-rows supports CSV only")
        import tempfile
        block = load_protocol_mapping(args.config, args.protocol)
        work_file = str(Path(tempfile.mkdtemp(prefix="cdd_clean_")) / Path(args.file).name)
        kept, dropped = write_cleaned_csv(args.file, block, work_file)
        print(f"drop-empty-rows: kept={kept} dropped={dropped}  cleaned_copy={work_file}\n")

    # Pre-flight: if a protocol is named, validate the header against config first.
    if args.protocol is not None:
        if not check_mapping(work_file, args.config, args.protocol):
            sys.exit("ERR: mapping check failed — fix UNMAPPED columns or re-run "
                     "with --check-mapping to inspect. Aborting before upload.")
        print()

    mapping_json = None
    if args.mapping_json:
        mapping_json = json.loads(Path(args.mapping_json).read_text())
        if not isinstance(mapping_json, dict):
            sys.exit("ERR: --mapping-json must contain a JSON object")

    payload = build_payload(
        project=project,
        mapping_template=mapping_template,
        mapping_json=mapping_json,
        autoreject=not args.no_autoreject,
        runs=_build_runs(args),
    )

    if args.dry_run:
        p = Path(work_file)
        print("=== dry-run (nothing posted) ===")
        print(f"file={p} exists={p.exists()} "
              f"bytes={p.stat().st_size if p.exists() else 'NA'}")
        print(f"payload={json.dumps(payload, indent=2)}")
        return

    slurp_id = submit_slurp(session, args.vault, work_file, payload)
    final = poll_slurp(session, args.vault, slurp_id,
                       interval=args.poll_interval, timeout=args.poll_timeout)
    ok = report_outcome(final)
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
