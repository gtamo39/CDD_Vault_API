#!/usr/bin/env python3
"""
convert_upload.py — Turn a raw WuXi ".xlsx" workbook into CDD-ready CSV rows.

Consolidates the previously ad-hoc raw->processed conversion (documented in
data/uploads/20260707/report.md) into one reusable module, plus an in-memory
version of the species split (cf. split_species.py, which operates on already
written `_upl.csv` files). Used by the import webapp and re-usable from the CLI.

HARD RULE (same as import_to_protocol): cell VALUES are never printed. Callers
get rows back in memory / written to disk; stdout stays metadata-only.

What it does to the workbook's "Upload" tab (found case/space-insensitively):
  * reconstruct '<x' / '>x' qualifiers stored as a number + Excel display
    format (openpyxl number_format) — pd.read_excel / SheetJS silently drop them
  * normalise real date cells to YYYY-MM-DD text (ints/strings pass through)
  * drop stray empty-header columns
  * drop wholly-blank rows and (given a protocol block) rows whose identifier
    column(s) are all blank — CDD rejects those
  * reconcile the identifier column name to what the CDD template expects
    (MDR1's template wants 'Batch Molecule-Batch ID' while WuXi ships
    'Molecule-Batch ID'); config-driven, not hard-coded to a PID
  * split by a 'Species' column into one unit per species (microsomal)
"""

import csv
import re
from datetime import date, datetime
from pathlib import Path

# The WuXi un-prefixed identifier and the prefixed name MDR1's template expects.
_ID_UNPREFIXED = "Molecule-Batch ID"
_ID_PREFIXED = "Batch Molecule-Batch ID"
SPECIES_HEADER = "species"


class UploadSheetError(Exception):
    """Raised when a workbook has no sheet named (case/space-insensitively) 'Upload'."""


def _norm(s):
    """Collapse whitespace runs (mirror of import_to_protocol._norm) so multi-line
    CRO headers match clean single-line config keys."""
    return " ".join(str(s).split()) if s is not None else ""


def _blank(v):
    """True if a cell is empty / None / whitespace-only."""
    return v is None or str(v).strip() == ""


def _cell_value(cell):
    """One openpyxl cell -> a CDD-ready scalar.

    Rebuilds a `<`/`>` qualifier from the cell's display format, renders real
    dates as ISO text, and passes everything else through unchanged.
    """
    val = cell.value
    fmt = cell.number_format or ""
    if isinstance(val, (int, float)) and not isinstance(val, bool) and isinstance(fmt, str):
        m = re.match(r"^\\([<>])", fmt)
        if m:
            dm = re.search(r"\.(0+)", fmt)
            decimals = len(dm.group(1)) if dm else 0
            return f"{m.group(1)} {val:.{decimals}f}"
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    return "" if val is None else val


def read_upload_sheet(path):
    """Read a workbook's 'Upload' tab into (header, rows).

    Header/rows are cleaned: stray empty-header columns dropped, qualifiers
    reconstructed, dates ISO-formatted, wholly-blank rows removed. Raises
    UploadSheetError if there is no 'Upload' sheet.
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    name = next((s for s in wb.sheetnames if s.strip().lower() == "upload"), None)
    if name is None:
        wb.close()
        raise UploadSheetError(
            f"no 'Upload' sheet in {Path(path).name}; found: {wb.sheetnames}"
        )
    grid = list(wb[name].iter_rows(values_only=False))
    wb.close()
    if not grid:
        return [], []
    keep = [i for i, c in enumerate(grid[0]) if not _blank(c.value)]
    header = [str(grid[0][i].value).lstrip("﻿").strip() for i in keep]
    rows = []
    for r in grid[1:]:
        row = [_cell_value(r[i]) if i < len(r) else "" for i in keep]
        if all(_blank(v) for v in row):
            continue
        rows.append(row)
    return header, rows


def reconcile_identifier(header, block):
    """Rename the file's identifier column to the name the config block expects.

    WuXi ships 'Molecule-Batch ID'; MDR1's CDD template maps 'Batch
    Molecule-Batch ID'. When the block asks for the prefixed name and the file
    only has the un-prefixed one, rename it. Returns (new_header, renames).
    """
    idents = list((block.get("identifiers") or {}).values())
    hdr_norm = {_norm(h): i for i, h in enumerate(header)}
    header = list(header)
    renames = {}
    for want in idents:
        if _norm(want) in hdr_norm:
            continue
        if want == _ID_PREFIXED and _norm(_ID_UNPREFIXED) in hdr_norm:
            i = hdr_norm[_norm(_ID_UNPREFIXED)]
            renames[header[i]] = want
            header[i] = want
    return header, renames


def drop_blank_identifier_rows(header, rows, block):
    """Drop rows whose identifier column(s) are all blank. Returns (kept, dropped)."""
    idents = list((block.get("identifiers") or {}).values())
    hdr_norm = [_norm(h) for h in header]
    idx = [hdr_norm.index(_norm(n)) for n in idents if _norm(n) in hdr_norm]
    if not idx:
        return rows, 0
    kept, dropped = [], 0
    for r in rows:
        if all(i >= len(r) or _blank(r[i]) for i in idx):
            dropped += 1
        else:
            kept.append(r)
    return kept, dropped


def split_species(header, rows):
    """Split rows by a 'Species' column into [(species, rows), ...], first-seen order.

    Returns None if there is no species column (i.e. not a species file).
    Rows with a blank species value match nothing and are dropped.
    """
    hdr_norm = [_norm(h).lower() for h in header]
    if SPECIES_HEADER not in hdr_norm:
        return None
    sc = hdr_norm.index(SPECIES_HEADER)
    species_of = lambda r: str(r[sc]).strip() if sc < len(r) and r[sc] is not None else ""
    order, seen = [], set()
    for r in rows:
        sp = species_of(r)
        if sp and sp not in seen:
            seen.add(sp)
            order.append(sp)
    return [(sp, [r for r in rows if species_of(r) == sp]) for sp in order]


def write_csv(path, header, rows):
    """Write header + rows as UTF-8-BOM CSV (the format CDD/Excel expect)."""
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(header)
        for r in rows:
            w.writerow(["" if v is None else v for v in r])


def main():
    """Smoke-test: report structure of a workbook's Upload tab (no cell values)."""
    import argparse

    ap = argparse.ArgumentParser(description="Inspect a WuXi .xlsx Upload tab (metadata only).")
    ap.add_argument("file", help="Raw .xlsx workbook with an 'Upload' tab.")
    args = ap.parse_args()

    header, rows = read_upload_sheet(args.file)
    print(f"file={Path(args.file).name}  columns={len(header)}  data_rows={len(rows)}")
    units = split_species(header, rows)
    if units is None:
        print("species_column=no")
    else:
        print(f"species_column=yes  species={[sp for sp, _ in units]}  "
              f"rows_per_species={[len(rs) for _, rs in units]}")


if __name__ == "__main__":
    main()
